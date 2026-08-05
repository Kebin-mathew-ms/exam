import csv
import io
import openpyxl
from typing import List, Dict, Tuple
from sqlalchemy.orm import Session
from decimal import Decimal

# ReportLab components for PDF exports
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from app.models.exam import Subject, QuestionCategory, DifficultyLevel, QuestionType, Question, QuestionOption

class ImportExportService:
    """Service to handle question imports (CSV/Excel) and exports (CSV/Excel/PDF)."""
    
    @staticmethod
    def parse_csv_questions(file_content: bytes, db: Session) -> dict:
        """Parse CSV questions, validate fields, and return import report stats."""
        text_stream = io.StringIO(file_content.decode('utf-8-sig', errors='ignore'))
        reader = csv.DictReader(text_stream)
        
        return ImportExportService._process_rows(reader, db)

    @staticmethod
    def parse_excel_questions(file_content: bytes, db: Session) -> dict:
        """Parse Excel questions, validate fields, and return import report stats."""
        wb = openpyxl.load_workbook(filename=io.BytesIO(file_content), read_only=True)
        sheet = wb.active
        
        # Read header row
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return {"success_count": 0, "failure_count": 1, "errors": [{"row": 1, "error": "Empty spreadsheet"}]}
        
        headers = [str(cell).strip().lower() if cell else "" for cell in rows[0]]
        
        row_dicts = []
        for i, row in enumerate(rows[1:], start=2):
            row_dict = {}
            for col_idx, cell in enumerate(row):
                if col_idx < len(headers) and headers[col_idx]:
                    row_dict[headers[col_idx]] = cell
            row_dicts.append(row_dict)
            
        return ImportExportService._process_rows(row_dicts, db)

    @staticmethod
    def _process_rows(rows, db: Session) -> dict:
        success_count = 0
        failure_count = 0
        errors = []
        
        # Fetch lookups to optimize database hits
        subjects = {s.subject_code.lower(): s.id for s in db.query(Subject).all()}
        categories = {c.name.lower(): c.id for c in db.query(QuestionCategory).all()}
        difficulties = {d.name.lower(): d.id for d in db.query(DifficultyLevel).all()}
        qtypes = {t.name.lower(): t.id for t in db.query(QuestionType).all()}
        
        for idx, row in enumerate(rows, start=2):
            try:
                title = row.get("title") or row.get("question_title")
                description = row.get("description") or row.get("question_description")
                subject_code = row.get("subject_code")
                category_name = row.get("category") or row.get("question_category")
                difficulty_name = row.get("difficulty")
                qtype_name = row.get("question_type") or row.get("type")
                marks_str = row.get("marks")
                neg_marks_str = row.get("negative_marks") or row.get("neg_marks")
                explanation = row.get("explanation")
                
                # Check required fields
                if not all([title, description, subject_code, category_name, difficulty_name, qtype_name]):
                    raise ValueError("Missing required fields (title, description, subject_code, category, difficulty, type)")
                
                # Lookup translations
                sub_id = subjects.get(str(subject_code).strip().lower())
                if not sub_id:
                    raise ValueError(f"Subject code '{subject_code}' not found in database lookup")
                
                cat_id = categories.get(str(category_name).strip().lower())
                if not cat_id:
                    raise ValueError(f"Category '{category_name}' not found in database lookup")
                
                diff_id = difficulties.get(str(difficulty_name).strip().lower())
                if not diff_id:
                    raise ValueError(f"Difficulty level '{difficulty_name}' not found in database lookup")
                
                type_id = qtypes.get(str(qtype_name).strip().lower())
                if not type_id:
                    raise ValueError(f"Question Type '{qtype_name}' not found in database lookup")

                # Parse marks
                marks = Decimal(str(marks_str or "1.00"))
                neg_marks = Decimal(str(neg_marks_str or "0.00"))
                
                # Create question
                db_question = Question(
                    title=str(title).strip(),
                    description=str(description).strip(),
                    subject_id=sub_id,
                    category_id=cat_id,
                    difficulty_id=diff_id,
                    question_type_id=type_id,
                    marks=marks,
                    negative_marks=neg_marks,
                    explanation=str(explanation).strip() if explanation else None,
                    status="active"
                )
                db.add(db_question)
                db.flush() # Populate ID for options linking
                
                # Option Parsing (MCQ Options: Option A;Option B;Option C)
                options_str = row.get("options")
                correct_str = row.get("correct_option") or row.get("correct_answer")
                
                if type_id == 1:  # MCQ
                    if not options_str:
                        raise ValueError("MCQ Question requires options column list (semicolon separated)")
                    
                    option_items = [opt.strip() for opt in str(options_str).split(";") if opt.strip()]
                    if len(option_items) < 2:
                        raise ValueError("MCQ must have at least two options")
                    
                    correct_idx = -1
                    if correct_str:
                        correct_val = str(correct_str).strip()
                        # If a string, match option text, otherwise check integer index
                        if correct_val.isdigit():
                            correct_idx = int(correct_val) - 1
                        else:
                            for c_i, opt_text in enumerate(option_items):
                                if opt_text.lower() == correct_val.lower():
                                    correct_idx = c_i
                                    break
                    
                    for o_idx, opt_text in enumerate(option_items):
                        db_option = QuestionOption(
                            question_id=db_question.id,
                            option_text=opt_text,
                            is_correct=(o_idx == correct_idx),
                            display_order=o_idx + 1
                        )
                        db.add(db_option)
                
                elif type_id == 2:  # True False
                    correct_bool = str(correct_str or "").strip().lower() in ("true", "t", "1", "yes")
                    db_options = [
                        QuestionOption(question_id=db_question.id, option_text="True", is_correct=correct_bool, display_order=1),
                        QuestionOption(question_id=db_question.id, option_text="False", is_correct=not correct_bool, display_order=2)
                    ]
                    db.add_all(db_options)

                db.commit()
                success_count += 1
            except Exception as e:
                db.rollback()
                failure_count += 1
                errors.append({"row": idx, "error": str(e)})
                
        return {
            "success_count": success_count,
            "failure_count": failure_count,
            "errors": errors
        }

    @staticmethod
    def export_questions_csv(questions: List[Question]) -> str:
        """Export list of questions to CSV text format."""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write Headers
        writer.writerow([
            "title", "description", "subject_code", "category", "difficulty", 
            "question_type", "marks", "negative_marks", "explanation", "options", "correct_option"
        ])
        
        for q in questions:
            # Options list serialization
            opts = ";".join([o.option_text for o in q.options])
            correct = next((o.option_text for o in q.options if o.is_correct), "")
            
            writer.writerow([
                q.title,
                q.description,
                q.subject.subject_code,
                q.category.name,
                q.difficulty.name,
                q.question_type.name,
                float(q.marks),
                float(q.negative_marks),
                q.explanation or "",
                opts,
                correct
            ])
            
        return output.getvalue()

    @staticmethod
    def export_questions_excel(questions: List[Question]) -> bytes:
        """Export list of questions to Excel (.xlsx) file buffer."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Question Bank"
        
        # Header Row Styling
        headers = [
            "Title", "Description", "Subject Code", "Category", "Difficulty", 
            "Question Type", "Marks", "Negative Marks", "Explanation", "Options (Semi-colon separated)", "Correct Answer"
        ]
        ws.append(headers)
        
        for q in questions:
            opts = ";".join([o.option_text for o in q.options])
            correct = next((o.option_text for o in q.options if o.is_correct), "")
            ws.append([
                q.title,
                q.description,
                q.subject.subject_code,
                q.category.name,
                q.difficulty.name,
                q.question_type.name,
                float(q.marks),
                float(q.negative_marks),
                q.explanation or "",
                opts,
                correct
            ])
            
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def export_questions_pdf(questions: List[Question]) -> bytes:
        """Export list of questions to formatted PDF document."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
        
        styles = getSampleStyleSheet()
        
        # Add custom paragraph styles safely
        title_style = ParagraphStyle(
            name='PDFDocTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#4F46E5'),
            spaceAfter=20
        )
        q_title_style = ParagraphStyle(
            name='PDFQuestionTitle',
            parent=styles['Heading3'],
            fontSize=11,
            spaceBefore=10,
            spaceAfter=5
        )
        body_style = ParagraphStyle(
            name='PDFBodyStyle',
            parent=styles['Normal'],
            fontSize=9,
            spaceAfter=3
        )
        
        story = []
        story.append(Paragraph("Aegis Online Examination System - Question Bank", title_style))
        story.append(Spacer(1, 10))
        
        for idx, q in enumerate(questions, start=1):
            story.append(Paragraph(f"<b>Q{idx}. {q.title}</b> [Type: {q.question_type.name} | Marks: {q.marks}]", q_title_style))
            desc_cleaned = q.description.replace("<p>", "").replace("</p>", "").replace("<strong>", "").replace("</strong>", "")
            story.append(Paragraph(desc_cleaned, body_style))
            
            # Print options if MCQ
            if q.options:
                opts_text = []
                for o in q.options:
                    mark = " [Correct]" if o.is_correct else ""
                    opts_text.append(f" - {o.option_text}{mark}")
                opts_paragraph = "<br/>".join(opts_text)
                story.append(Paragraph(opts_paragraph, body_style))
                
            if q.explanation:
                story.append(Paragraph(f"<i>Explanation: {q.explanation}</i>", body_style))
                
            story.append(Spacer(1, 12))
            
        doc.build(story)
        return buffer.getvalue()
